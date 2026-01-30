from __future__ import annotations
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
import openpyxl

@dataclass(frozen=True)
class GlossaryEntry:
    cn: str
    en: str
    priority: int = 0  # 预留：越大越优先

def _norm(s: str) -> str:
    return s.strip()

def load_glossary_xlsx(
    path: str,
    sheet_name: Optional[str] = None,
    cn_col: str = "CN",
    en_col: str = "EN",
) -> List[GlossaryEntry]:
    """
    从 xlsx 读取术语表。默认寻找表头 CN/EN。
    你也可以改 cn_col/en_col 适配你的表头。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # 读取表头行（默认第一行）
    headers: Dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        headers[str(cell.value).strip()] = idx

    if cn_col not in headers or en_col not in headers:
        raise ValueError(
            f"Glossary headers not found. Need columns '{cn_col}' and '{en_col}'. "
            f"Found: {list(headers.keys())}"
        )

    cn_idx = headers[cn_col]
    en_idx = headers[en_col]

    entries: List[GlossaryEntry] = []
    for r in range(2, ws.max_row + 1):
        cn = ws.cell(r, cn_idx).value
        en = ws.cell(r, en_idx).value
        if cn is None or en is None:
            continue
        cn_s = _norm(str(cn))
        en_s = _norm(str(en))
        if not cn_s or not en_s:
            continue
        entries.append(GlossaryEntry(cn=cn_s, en=en_s))

    # 关键：按中文长度降序排序，避免“导柱”抢先替换“导柱组件”
    entries.sort(key=lambda x: (len(x.cn), x.priority), reverse=True)
    return entries
