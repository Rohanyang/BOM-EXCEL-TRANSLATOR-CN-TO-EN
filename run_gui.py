# run_gui.py
# Company-friendly BOM Translator GUI (Excel -> Excel) using WORD LIST.xlsx
# - Works for both .py run and PyInstaller --onefile .exe
# - Reads glossary from exe directory (next to exe) by default
# - Translates by whole-cell exact match (prevents substring overlap bugs)
# - Outputs *_EN.xlsx and *_QA_untranslated.csv next to input file

import sys
import csv
import re
from pathlib import Path
from typing import Dict, List, Tuple, Set

import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import load_workbook


APP_TITLE = "MLD BOM Translator"
GLOSSARY_FILENAME = "WORD LIST.xlsx"

# Glossary columns (your WORD LIST headers)
CN_HEADER = "CN"
EN_HEADER = "EN"


def get_app_dir() -> Path:
    """
    When packed by PyInstaller --onefile:
      sys.executable = path to the extracted runner exe location (actual exe path)
      Using parent of sys.executable ensures we use the directory where the exe sits.
    When running as .py:
      __file__ points to this script.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def find_header_row(ws, required_headers: Tuple[str, str]) -> Tuple[int, Dict[str, int]]:
    """
    Find a row that contains both headers (CN/EN) and return:
    (row_index_1based, {header_name: column_index_1based})
    """
    cn_h, en_h = required_headers
    for r in range(1, min(ws.max_row, 50) + 1):  # scan first 50 rows
        header_map = {}
        for c in range(1, min(ws.max_column, 50) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                vv = v.strip()
                if vv == cn_h:
                    header_map[cn_h] = c
                elif vv == en_h:
                    header_map[en_h] = c
        if cn_h in header_map and en_h in header_map:
            return r, header_map
    raise ValueError(f"未找到表头：{cn_h}/{en_h}（请确认 WORD LIST.xlsx 中存在 CN 和 EN 两列）")


def load_glossary_xlsx(path: Path, sheet_name: str | None = None) -> Dict[str, str]:
    """
    Load CN->EN mapping from WORD LIST.xlsx.
    - Requires headers CN and EN.
    - CN must be unique; duplicates will be warned and last one wins.
    """
    if not path.exists():
        raise FileNotFoundError(str(path))

    wb = load_workbook(filename=str(path), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row, header_map = find_header_row(ws, (CN_HEADER, EN_HEADER))
    cn_col = header_map[CN_HEADER]
    en_col = header_map[EN_HEADER]

    mapping: Dict[str, str] = {}
    dup_cn: Set[str] = set()

    for r in range(header_row + 1, ws.max_row + 1):
        cn_val = ws.cell(row=r, column=cn_col).value
        en_val = ws.cell(row=r, column=en_col).value

        if cn_val is None:
            continue

        if not isinstance(cn_val, str):
            cn_key = str(cn_val).strip()
        else:
            cn_key = cn_val.strip()

        if not cn_key:
            continue

        en_str = ""
        if en_val is None:
            en_str = ""
        elif isinstance(en_val, str):
            en_str = en_val.strip()
        else:
            en_str = str(en_val).strip()

        if cn_key in mapping:
            dup_cn.add(cn_key)

        mapping[cn_key] = en_str

    if dup_cn:
        # Not fatal: show warning
        sample = ", ".join(list(sorted(dup_cn))[:10])
        messagebox.showwarning(
            "术语表提示",
            f"WORD LIST 中发现重复 CN（将以最后一条为准）。示例：\n{sample}\n\n建议你们清理重复词条以保证一致性。"
        )

    return mapping


def contains_chinese(s: str) -> bool:
    # Basic CJK detection
    return any("\u4e00" <= ch <= "\u9fff" for ch in s)



def _normalize_text(s: str) -> str:
    # 基础清洗：去首尾空格、统一全角空格、去掉多余换行两侧空格
    s = s.replace("\u3000", " ")  # 全角空格 -> 半角
    s = re.sub(r"\s+", " ", s)    # 多个空白 -> 单个空格（含换行/tab）
    return s.strip()

def _build_longest_first_terms(glossary: dict) -> list[tuple[str, str]]:
    # 只拿中文key（避免英文再命中），按长度倒序，保证最长词优先
    items = []
    for k, v in glossary.items():
        if not k or not v:
            continue
        # 只处理含中文的key，避免数字/符号 key 造成误伤
        if any("\u4e00" <= ch <= "\u9fff" for ch in k):
            items.append((k, v))
    items.sort(key=lambda x: len(x[0]), reverse=True)
    return items

def _replace_mixed_text(src: str, terms: list[tuple[str, str]]) -> str:
    """
    安全混合替换：最长词优先、非重叠式。
    因为中文->英文，替换后不会再次命中中文key，所以不会连环替换。
    """
    out = src
    for cn, en in terms:
        if cn in out:
            out = out.replace(cn, en)
    return out

def translate_workbook_exact(in_path, out_path, glossary):
    wb = load_workbook(filename=str(in_path))
    untranslated = []

    terms = _build_longest_first_terms(glossary)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue

                raw = v
                src = _normalize_text(raw)
                if not src:
                    continue

                # 1) 整格精确匹配（最稳）
                if src in glossary and glossary[src]:
                    cell.value = glossary[src]
                    continue

                # 2) 混合文本替换（最长词优先）
                mixed = _replace_mixed_text(src, terms)

                final_text = mixed if mixed != src else src

                # 写回（如果有变化才写）
                if final_text != raw:
                    cell.value = final_text
                
                # ✅ QA 判定：只要最终内容仍含中文，就进 QA
                if any("\u4e00" <= ch <= "\u9fff" for ch in final_text):
                    # 提取仍残留的中文片段（方便直接复制到 WORD LIST 的 CN 列）
                    cn_chunks = " | ".join(sorted(set(re.findall(r"[\u4e00-\u9fff]+", final_text))))

                    # 记录：sheet, row, cell, raw_original, after_translate, cn_chunks
                    untranslated.append((
                        ws.title,
                        cell.row,
                        cell.coordinate,
                        raw,         # 原始单元格内容（未清洗）
                        final_text,  # 替换后最终内容（可能是部分翻译）
                        cn_chunks    # 剩余中文片段
                    ))

                
                    

    wb.save(str(out_path))
    return untranslated



def write_untranslated_csv(csv_path: Path, records) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "row", "cell", "raw_original", "after_translate", "cn_chunks"])
        for sheet, row, cell, raw_original, after_translate, cn_chunks in records:
            w.writerow([sheet, row, cell, raw_original, after_translate, cn_chunks])



def build_output_paths(in_file: Path) -> Tuple[Path, Path]:
    # output next to input: xxx_EN.xlsx
    out_xlsx = in_file.with_name(f"{in_file.stem}_EN{in_file.suffix}")
    qa_csv = in_file.with_name(f"{in_file.stem}_QA_untranslated.csv")
    return out_xlsx, qa_csv


def main():
    app_dir = get_app_dir()
    glossary_path = app_dir / GLOSSARY_FILENAME

    # --- GUI root ---
    root = tk.Tk()
    root.title(APP_TITLE)
    root.geometry("520x260")
    root.resizable(False, False)

    # --- Handlers ---
    def pick_and_run():
        # ensure glossary exists next to exe
        if not glossary_path.exists():
            messagebox.showerror(
                "缺少术语表",
                f"未找到术语表：\n{glossary_path}\n\n请把 {GLOSSARY_FILENAME} 放在 exe 同目录（与程序放一起）。"
            )
            return

        try:
            glossary = load_glossary_xlsx(glossary_path, sheet_name=None)
        except Exception as e:
            messagebox.showerror("术语表读取失败", str(e))
            return

        files = filedialog.askopenfilenames(
            title="选择要翻译的 BOM Excel（可多选）",
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if not files:
            return

        ok_count = 0
        fail_list: List[str] = []

        for fp in files:
            in_path = Path(fp)
            try:
                out_xlsx, qa_csv = build_output_paths(in_path)

                untranslated = translate_workbook_exact(
                    in_path=in_path,
                    out_path=out_xlsx,
                    glossary=glossary,
                )

                # Only write QA file if there are untranslated items
                if untranslated:
                    write_untranslated_csv(qa_csv, untranslated)
                else:
                    # if no untranslated, remove old QA if exists
                    if qa_csv.exists():
                        try:
                            qa_csv.unlink()
                        except Exception:
                            pass

                ok_count += 1
            except Exception as e:
                fail_list.append(f"{in_path.name} -> {e}")

        if fail_list:
            messagebox.showwarning(
                "部分文件失败",
                f"成功：{ok_count} 个\n失败：{len(fail_list)} 个\n\n" + "\n".join(fail_list[:8])
            )
        else:
            messagebox.showinfo("完成", f"翻译完成：{ok_count} 个文件\n输出文件与原文件同目录，带 _EN 后缀。")

    # --- UI ---
    label = tk.Label(
        root,
        text=f"词库路径（同目录读取）：\n{glossary_path}",
        justify="left",
        wraplength=500
    )
    label.pack(pady=16)

    btn = tk.Button(root, text="选择 BOM 文件并翻译（可多选）", width=36, height=2, command=pick_and_run)
    btn.pack(pady=12)

    tip = tk.Label(
        root,
        text="输出：原文件名_EN.xlsx\n如有未翻译中文，会生成：*_QA_untranslated.csv（用于补充词库）",
        justify="left"
    )
    tip.pack(pady=10)

    root.mainloop()


if __name__ == "__main__":
    main()
