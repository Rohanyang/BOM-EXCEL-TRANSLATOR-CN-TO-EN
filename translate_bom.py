from __future__ import annotations
import argparse
from pathlib import Path
from typing import List, Dict, Any, Optional, Set

import openpyxl
from tqdm import tqdm

from glossary import load_glossary_xlsx, GlossaryEntry
from rules import should_skip_cell, has_chinese
from qa import write_qa_csv

def replace_terms(text: str, entries: List[GlossaryEntry]) -> str:
    """按术语表做子串替换（长度降序已在 glossary 处理）。"""
    s = str(text)
    for e in entries:
        if e.cn in s:
            s = s.replace(e.cn, e.en)
    return s

def translate_workbook(
    in_xlsx: str,
    out_xlsx: str,
    glossary_entries: List[GlossaryEntry],
    only_sheets: Optional[Set[str]] = None,
    only_columns: Optional[Set[str]] = None,
) -> List[Dict[str, Any]]:
    """
    only_columns: 例如 {"C","D","F"} 只翻译这些列（强烈建议：只翻译“品名/备注”等列）
    返回：QA 未翻译中文列表
    """
    wb = openpyxl.load_workbook(in_xlsx)  # 不用 data_only，保留公式（我们只改文本）
    qa_rows: List[Dict[str, Any]] = []

    sheet_names = wb.sheetnames
    for sname in sheet_names:
        if only_sheets and sname not in only_sheets:
            continue
        ws = wb[sname]

        # tqdm 进度条按行
        for row in tqdm(range(1, ws.max_row + 1), desc=f"Sheet: {sname}", leave=False):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                v = cell.value

                if v is None:
                    continue
                if not isinstance(v, str):
                    continue  # 只翻译字符串，数字/日期/公式等不动

                # 仅翻译指定列（推荐）
                if only_columns:
                    col_letter = openpyxl.utils.get_column_letter(col)
                    if col_letter not in only_columns:
                        continue

                if should_skip_cell(v):
                    continue

                new_v = replace_terms(v, glossary_entries)
                if new_v != v:
                    cell.value = new_v

        # 做一次 QA：扫描 sheet 内仍含中文的字符串单元格（同样受 only_columns 限制）
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                v = cell.value
                if v is None or not isinstance(v, str):
                    continue
                if only_columns:
                    col_letter = openpyxl.utils.get_column_letter(col)
                    if col_letter not in only_columns:
                        continue
                if has_chinese(v):
                    addr = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    qa_rows.append({
                        "sheet": sname,
                        "row": row,
                        "col": openpyxl.utils.get_column_letter(col),
                        "cell": addr,
                        "original": v
                    })

    wb.save(out_xlsx)
    return qa_rows

def main():
    p = argparse.ArgumentParser(description="BOM glossary-based translator (CN->EN) for Excel.")
    p.add_argument("--in", dest="in_xlsx", required=True, help="Input BOM .xlsx")
    p.add_argument("--out", dest="out_xlsx", required=True, help="Output translated .xlsx")
    p.add_argument("--glossary", required=True, help="Glossary .xlsx path")
    p.add_argument("--glossary-sheet", default=None, help="Glossary sheet name (optional)")
    p.add_argument("--cn-col", default="CN", help="Glossary CN header name (default CN)")
    p.add_argument("--en-col", default="EN", help="Glossary EN header name (default EN)")
    p.add_argument("--only-sheets", default=None, help="Comma separated sheet names to translate")
    p.add_argument("--only-columns", default=None, help="Comma separated Excel columns (e.g. C,D,F)")
    p.add_argument("--qa", default=None, help="Output QA csv path (optional)")

    args = p.parse_args()

    entries = load_glossary_xlsx(
        path=args.glossary,
        sheet_name=args.glossary_sheet,
        cn_col=args.cn_col,
        en_col=args.en_col,
    )

    only_sheets = set([s.strip() for s in args.only_sheets.split(",") if s.strip()]) if args.only_sheets else None
    only_columns = set([c.strip().upper() for c in args.only_columns.split(",") if c.strip()]) if args.only_columns else None

    qa_rows = translate_workbook(
        in_xlsx=args.in_xlsx,
        out_xlsx=args.out_xlsx,
        glossary_entries=entries,
        only_sheets=only_sheets,
        only_columns=only_columns
    )

    qa_path = args.qa or (str(Path(args.out_xlsx).with_suffix("")) + "_QA_untranslated.csv")
    write_qa_csv(qa_path, qa_rows)

    print(f"Done.\n- Output: {args.out_xlsx}\n- QA: {qa_path}\n- Untranslated cells: {len(qa_rows)}")

if __name__ == "__main__":
    main()
